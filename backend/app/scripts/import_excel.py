"""
Importa Cuentas_boxphone.xlsx tal cual, de la forma más fiel posible a los
datos originales.

Uso (desde backend/, con el venv activo y las variables de entorno puestas
igual que para correr uvicorn):

    python -m app.scripts.import_excel /ruta/a/Cuentas_boxphone.xlsx

Reglas aplicadas (decididas junto con Álvaro):
- Humanización vive en cada red social individual, no en la cuenta.
- Redes marcadas "Sí" en Correos pero sin usuario/contraseña -> se crean como
  "pendiente de credenciales" (username/password en NULL).
- Celular ambiguo o sin identificar en 'Correos' -> se importa sin celular
  asignado, guardando el texto original en las notas.
- Los identificadores de celular "PC 5".."PC 20" en la hoja 'Facebook' son
  el MISMO celular que 'CEL 5'..'CEL 20' (confirmado con Álvaro) — se
  normalizan a CEL N. La fuente de verdad del nickname de cada celular es
  la hoja 'Resumen Facebook'.
- El cruce entre 'Correos' (persona/correo) y 'Facebook' (login real de esa
  red) se hace por nombre exacto (sin mayúsculas/espacios). Si el nombre no
  matchea con exactamente una persona en 'Correos', NO se inventa una cuenta
  nueva (no hay correo válido de dónde partir) — se reporta al final para
  revisión manual.
- Solo 4 cuentas traen código 2FA (excepción heredada); el resto no.
"""
import re
import sys
from datetime import datetime

import openpyxl

from app.core.database import SessionLocal
from app.core.crypto import encrypt
from app.core.enums import ProxyStatus, Status
from app.models.proxy import Proxy
from app.models.device import Device
from app.models.account import Account
from app.models.social_account import SocialAccount

CEL_RE = re.compile(r"^(?:CEL|PC)\s*0*([0-9]{1,2})$", re.IGNORECASE)
ID_IN_URL_RE = re.compile(r"id=(\d+)")


def parse_cel_number(raw) -> int | None:
    """Devuelve el número de celular SOLO si el texto es exactamente
    'CEL N' o 'PC N' (ignorando may/min y espacios). None si es ambiguo."""
    if raw is None:
        return None
    text = str(raw).strip()
    m = CEL_RE.fullmatch(text)
    return int(m.group(1)) if m else None


def truthy_si(raw) -> bool:
    return bool(raw) and str(raw).strip().lower() in ("si", "sí")


def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)


# ---------------------------------------------------------------------------
# 1) Nicknames de celular (fuente de verdad: Resumen Facebook)
# ---------------------------------------------------------------------------

def build_nickname_map(wb) -> dict[int, str]:
    ws = wb["Resumen Facebook"]
    out = {}
    for row in ws.iter_rows(min_row=2, max_row=21, values_only=True):
        num = parse_cel_number(row[1])
        if num and row[2]:
            out[num] = str(row[2]).strip()
    return out


# ---------------------------------------------------------------------------
# 2) Proxys (uno por celular, string combinado ip:puerto:usuario:pass)
# ---------------------------------------------------------------------------

def build_proxy_map(wb) -> dict[int, dict]:
    ws = wb["proxys"]
    out = {}
    for row in ws.iter_rows(min_row=2, max_row=1000, values_only=True):
        num = parse_cel_number(row[0])
        if not num:
            continue
        proxy_str = row[4]
        status_txt = (row[5] or "").strip().lower()
        status = ProxyStatus.inoperativo if "remplazar" in status_txt else ProxyStatus.operativo
        parts = str(proxy_str).split(":") if proxy_str else []
        if len(parts) != 4:
            print(f"  [!] CEL {num}: proxy con formato inesperado ({proxy_str!r}), se omite")
            continue
        ip, port, user, pw = parts
        out[num] = {"ip": ip, "port": int(port), "username": user,
                    "password": pw, "status": status}
    return out


# ---------------------------------------------------------------------------
# 3) Perfiles y link -> nombre (may/min) -> profile_url, e id numérico -> nombre
# ---------------------------------------------------------------------------

def build_profile_link_maps(wb):
    ws = wb["Perfiles y link"]
    name_to_url = {}
    id_to_name = {}
    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
        # filas "CEL N" son encabezados de grupo (col B y C vacías); las de
        # datos son (orden, nombre, url)
        if row[1] is None:
            continue
        nombre, url = row[1], row[2]
        if nombre:
            name_to_url[str(nombre).strip().lower()] = url
        if url:
            m = ID_IN_URL_RE.search(str(url))
            if m and nombre:
                id_to_name[m.group(1)] = str(nombre).strip()
    return name_to_url, id_to_name


# ---------------------------------------------------------------------------
# 4) Devices (1..20) con nickname + proxy
# ---------------------------------------------------------------------------

def create_devices(db, nickname_map, proxy_map) -> dict[int, Device]:
    devices = {}
    for num in range(1, 21):
        proxy = None
        pdata = proxy_map.get(num)
        if pdata:
            proxy = Proxy(ip=pdata["ip"], port=pdata["port"], username=pdata["username"],
                          password_encrypted=encrypt(pdata["password"]), status=pdata["status"])
            db.add(proxy)
            db.flush()
        device = Device(name=f"CEL {num}", nickname=nickname_map.get(num),
                        proxy_id=proxy.id if proxy else None)
        db.add(device)
        db.flush()
        devices[num] = device
    print(f"  {len(devices)} celulares creados (CEL 1..CEL 20) con sus proxys.")
    return devices


# ---------------------------------------------------------------------------
# 5) Correos -> Account (persona) + socials "pendiente" según banderas Sí
# ---------------------------------------------------------------------------

def resolve_device_ref(raw, devices, nickname_map):
    """Devuelve (device_or_None, nota_o_None) según texto libre de celular."""
    if raw is None or str(raw).strip() == "":
        return None, None
    text = str(raw).strip()
    num = parse_cel_number(text)
    if num and num in devices:
        return devices[num], None
    # ¿matchea un nickname exacto (ej. "9KL")?
    upper = text.upper()
    for n, nick in nickname_map.items():
        if nick.upper() == upper and n in devices:
            return devices[n], None
    return None, f"Celular original en Excel: '{text}' (ambiguo, revisar manualmente)"


def import_correos(db, wb, devices, nickname_map) -> dict[str, Account]:
    ws = wb["Correos"]
    accounts_by_email: dict[str, Account] = {}
    seen_emails: set[str] = set()
    dupe_count = 0

    for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
        email_raw = row[0]
        if not email_raw:
            continue
        email = email_raw.strip()
        nombre = (row[1] or "").strip() or None
        num_carpeta_raw = row[2]
        fecha_nac = row[3]
        contrasena_redes = row[4] or ""
        cel_raw = row[5]
        fb_flag, ig_flag, tt_flag = row[6], row[7], row[8]

        # duplicado de correo: no se descarta, se hace único con sufijo
        notes_parts = []
        key = email.lower()
        if key in seen_emails:
            dupe_count += 1
            original = email
            email = email.split("@")[0] + f"+dup{dupe_count}@" + email.split("@")[1]
            notes_parts.append(
                f"Correo duplicado en el Excel (original: {original}); "
                f"revisar cuál de las dos personas es la correcta.")
        seen_emails.add(key)

        sequence_number = None
        if isinstance(num_carpeta_raw, (int, float)):
            sequence_number = int(num_carpeta_raw)
        elif num_carpeta_raw:
            notes_parts.append(f"Nota original (columna 'Numero de Carpeta'): {num_carpeta_raw}")

        birth_date = fecha_nac.date() if isinstance(fecha_nac, datetime) else None

        device, dev_note = resolve_device_ref(cel_raw, devices, nickname_map)
        if dev_note:
            notes_parts.append(dev_note)

        account = Account(
            corporate_email=email,
            corp_password_encrypted=encrypt(contrasena_redes),
            status=Status.activo,
            profile_name=nombre,
            birth_date=birth_date,
            sequence_number=sequence_number,
            device_id=device.id if device else None,
            notes="; ".join(notes_parts) if notes_parts else None,
        )
        db.add(account)
        db.flush()

        for platform, flag in (("facebook", fb_flag), ("instagram", ig_flag), ("tiktok", tt_flag)):
            if truthy_si(flag):
                account.social_accounts.append(SocialAccount(platform=platform))

        accounts_by_email[email.lower()] = account
        if nombre:
            accounts_by_email.setdefault(f"name::{nombre.strip().lower()}", account)

    print(f"  {len(seen_emails)} correos procesados, {dupe_count} duplicado(s) detectado(s).")
    return accounts_by_email


# ---------------------------------------------------------------------------
# 6) Facebook (detalle real) -> enriquece/crea el SocialAccount facebook
# ---------------------------------------------------------------------------

def parse_facebook_row(row):
    cel, nombre, perfil, usuario, pw, fa2, combinado, datos, estado = row
    nombre = (nombre or "").strip() or None
    slot = int(perfil) if perfil else None

    if not (nombre or usuario or combinado):
        return None  # slot vacío, nada que importar

    if usuario and pw:
        username, password, profile_id = str(usuario), pw, None
    elif combinado:
        parts = str(combinado).split(":")
        parts = [p.strip() for p in parts if p != ""]
        if len(parts) >= 2:
            username, password = parts[0], parts[1]
            profile_id = parts[3] if len(parts) > 3 else None
        else:
            return None
    else:
        return None

    return {
        "device_num": parse_cel_number(cel), "nombre": nombre, "slot": slot,
        "username": username, "password": password,
        "profile_id": profile_id, "nota": datos,
    }


def import_facebook_detail(db, wb, accounts_by_email, name_to_url, id_to_name):
    ws = wb["Facebook"]
    sin_match = []
    aplicados = 0

    for raw_row in ws.iter_rows(min_row=2, max_row=1000, values_only=True):
        parsed = parse_facebook_row(raw_row)
        if not parsed:
            continue

        nombre = parsed["nombre"]
        if not nombre:
            candidate_id = parsed["profile_id"] or parsed["username"]
            if candidate_id and str(candidate_id).strip() in id_to_name:
                nombre = id_to_name[str(candidate_id).strip()]

        account = accounts_by_email.get(f"name::{(nombre or '').strip().lower()}") if nombre else None
        if not account:
            sin_match.append({"celular": f"CEL {parsed['device_num']}", "nombre": nombre,
                              "usuario": parsed["username"]})
            continue

        profile_url = name_to_url.get((nombre or "").strip().lower())

        existing_fb = next((s for s in account.social_accounts if s.platform.value == "facebook"), None)
        if existing_fb is None:
            existing_fb = SocialAccount(platform="facebook")
            account.social_accounts.append(existing_fb)

        existing_fb.username = parsed["username"]
        existing_fb.password_encrypted = encrypt(parsed["password"])
        existing_fb.slot_number = parsed["slot"]
        existing_fb.profile_url = profile_url
        if parsed["nota"]:
            existing_fb.notes = parsed["nota"]

        # si el celular del detalle FB difiere del que ya tenía la cuenta,
        # no lo pisamos: solo lo anotamos para revisión manual
        if parsed["device_num"] and account.device and account.device.name != f"CEL {parsed['device_num']}":
            note = (f"Aviso: hoja 'Facebook' ubica esta cuenta en CEL {parsed['device_num']}, "
                   f"pero 'Correos' la tenía en {account.device.name}. Revisar manualmente.")
            account.notes = f"{account.notes}; {note}" if account.notes else note

        aplicados += 1

    print(f"  {aplicados} cuenta(s) de Facebook con credenciales reales aplicadas.")
    if sin_match:
        print(f"  {len(sin_match)} fila(s) de 'Facebook' SIN cuenta madre en 'Correos' (no importadas, revisar a mano):")
        for s in sin_match:
            print(f"    - {s['celular']} | nombre: {s['nombre']!r} | usuario: {s['usuario']!r}")


# ---------------------------------------------------------------------------

def main(path: str):
    print(f"Importando {path} ...")
    wb = load_workbook(path)
    db = SessionLocal()
    try:
        print("1) Nicknames de celular (Resumen Facebook)...")
        nickname_map = build_nickname_map(wb)

        print("2) Proxys (proxys)...")
        proxy_map = build_proxy_map(wb)

        print("3) Links de perfil (Perfiles y link)...")
        name_to_url, id_to_name = build_profile_link_maps(wb)

        print("4) Creando celulares...")
        devices = create_devices(db, nickname_map, proxy_map)

        print("5) Importando correos/personas (Correos)...")
        accounts_by_email = import_correos(db, wb, devices, nickname_map)

        print("6) Aplicando detalle real de Facebook (Facebook)...")
        import_facebook_detail(db, wb, accounts_by_email, name_to_url, id_to_name)

        db.commit()
        print("\nImportación completada y guardada.")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python -m app.scripts.import_excel /ruta/a/Cuentas_boxphone.xlsx")
        sys.exit(1)
    main(sys.argv[1])
